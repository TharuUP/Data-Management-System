from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import ComplaintForm
from .forms import ClassChangeForm
from .models import Complaint, ClassChange
from django.contrib.auth.decorators import login_required
from django_filters.views import FilterView
from .filters import ComplaintFilter, ClassChangeFilter
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.shortcuts import redirect
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.shortcuts import redirect
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import Group
from django.core.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404
from .forms import ClassChangeForm
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.template.loader import render_to_string
from weasyprint import HTML
from django.utils import timezone
from .models import Complaint
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import Workbook
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from .models import ClassChange  # Make sure it's ClassChange, not Complaint
from .forms import ClassChangeForm
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import ClassChange
from .forms import ClassChangeForm
from django.shortcuts import get_object_or_404, render, redirect
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.core.mail import EmailMessage


def custom_login_redirect(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            if user.is_superuser:
                return redirect('/admin/')  # Django admin for superusers

            # Check group and redirect to dashboard
            return redirect('dashboard')
    else:
        form = AuthenticationForm()

    return render(request, 'registration/login.html', {'form': form})

@login_required
def redirect_after_login(request):
    if request.user.is_superuser:
        return redirect('/admin/')
    else:
        return redirect('dashboard')

@login_required
def dashboard(request):
    return render(request, 'reports/dashboard.html')

@login_required
def add_class_change(request):
    if request.method == 'POST':
        form = ClassChangeForm(request.POST)
        if form.is_valid():
            change = form.save(commit=False)
            change.updated_by = request.user
            change.save()
            return redirect('dashboard')
    else:
        form = ClassChangeForm()
    return render(request, 'reports/add_class_change.html', {'form': form})


@login_required
def add_complaint(request):
    if request.method == 'POST':
        form = ComplaintForm(request.POST)
        if form.is_valid():
            complaint = form.save(commit=False)
            complaint.reporter = request.user
            complaint.save()

            # Prepare email content
            subject = f"New Complaint Registered: {complaint.complaint_no}"
            html_message = render_to_string('complaints/complaint_email.html', {'complaint': complaint})
            plain_message = strip_tags(html_message)
            recipient_email = complaint.area_manager_email
            
            send_mail(
                subject,
                plain_message,
                'your_email@gmail.com',  # Replace with your actual email configured in settings
                [recipient_email],
                html_message=html_message,
                fail_silently=False,
            )

            return redirect('dashboard')
    else:
        form = ComplaintForm()
    return render(request, 'reports/add_complaint.html', {'form': form})

@login_required
def complaint_list(request):
    complaint_filter = ComplaintFilter(request.GET, queryset=Complaint.objects.all())
    return render(request, 'reports/complaint_list.html', {'filter': complaint_filter})

@login_required
def class_change_list(request):
    class_filter = ClassChangeFilter(request.GET, queryset=ClassChange.objects.all())
    return render(request, 'reports/class_change_list.html', {'filter': class_filter})

def export_complaints_pdf(request):
    complaints = Complaint.objects.all()
    template_path = 'reports/complaint_pdf.html'
    context = {'complaints': complaints}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="complaints.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa.CreatePDF(html, dest=response)
    return response


@login_required
def edit_complaint(request, pk):
    complaint = get_object_or_404(Complaint, pk=pk)

    if not (complaint.reporter == request.user or request.user.is_superuser):
        raise PermissionDenied()

    old_email = complaint.area_manager_email

    if request.method == "POST":
        form = ComplaintForm(request.POST, instance=complaint)
        if form.is_valid():
            updated_complaint = form.save()

            new_email = updated_complaint.area_manager_email

            # If email changed, send apology and updated mail
            if old_email != new_email:
                send_mail(
                    'Correction: Previous Complaint Email Was Mistaken',
                    'Please ignore the previous complaint email sent in error.',
                    settings.DEFAULT_FROM_EMAIL,
                    [old_email],
                    fail_silently=False,
                )
                send_mail(
                    f'Updated Complaint Details: {updated_complaint.complaint_no}',
                    'Your complaint email was updated. Please check the details.',
                    settings.DEFAULT_FROM_EMAIL,
                    [new_email],
                    fail_silently=False,
                )

            # Prepare context for email template
            context = {'complaint': updated_complaint}

            # Render HTML email content using your template
            email_html_message = render_to_string('complaints/complaint_email.html', context)


            subject = f"Updated Complaint Details: {updated_complaint.complaint_no}"
            from_email = settings.DEFAULT_FROM_EMAIL
            to_email = [new_email]

            email = EmailMessage(subject, email_html_message, from_email, to_email)
            email.content_subtype = 'html'  # important for HTML email
            email.send()

            return redirect('complaint_list')

    else:
        form = ComplaintForm(instance=complaint)

    return render(request, 'reports/edit_complaint.html', {'form': form})

@login_required
def edit_class_change(request, pk):
    # Get the object or return 404 if not found
    obj = get_object_or_404(ClassChange, pk=pk)

    # Permission check:
    # Only the creator (updated_by) or admin (superuser) can edit
    if not (obj.updated_by == request.user or request.user.is_superuser):
        raise PermissionDenied()

    if request.method == 'POST':
        form = ClassChangeForm(request.POST, instance=obj)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.updated_by = request.user
            instance.save()
            return redirect('class_change_list')  # or wherever you want
    else:
        form = ClassChangeForm(instance=obj)

    return render(request, 'reports/edit_class_change.html', {'form': form})



def delete_complaint(request, pk):
    complaint = get_object_or_404(Complaint, pk=pk)

    # Access control check
    if not (complaint.reporter == request.user or request.user.is_superuser):
        raise PermissionDenied()

    complaint.delete()
    return redirect('complaint_list')  # or whatever your list view name is

def delete_class_change(request, pk):
    obj = get_object_or_404(ClassChange, pk=pk)

# Permission check:
    # Only the creator (updated_by) or admin (superuser) can edit
    if not (obj.updated_by == request.user or request.user.is_superuser):
        raise PermissionDenied()
        
    obj.delete()
    return redirect('class_change_list')


def export_classchange_pdf(request):
    class_changes = ClassChange.objects.all()
    template_path = 'reports/classchange_pdf.html'  # create this template for PDF rendering
    context = {'class_changes': class_changes}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="class_changes.pdf"'
    template = get_template(template_path)
    html = template.render(context)
    pisa.CreatePDF(html, dest=response)
    return response

@login_required
def some_report_view(request):
    queryset = Complaint.objects.all()

    total_count = queryset.count()
    pending_count = queryset.filter(status='Pending').count()
    done_count = queryset.filter(status='Done').count()
    progress_count = queryset.filter(status='In Progress').count()

    # Calculate percentages safely
    def percent(part, whole):
        return (part / whole * 100) if whole else 0

    pending_percent = percent(pending_count, total_count)
    done_percent = percent(done_count, total_count)
    progress_percent = percent(progress_count, total_count)

    context = {
        'complaints': queryset,
        'export_time': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_complaints': total_count,
        'pending_count': pending_count,
        'pending_percent': pending_percent,
        'done_count': done_count,
        'done_percent': done_percent,
        'progress_count': progress_count,
        'progress_percent': progress_percent,
    }

    return render(request, 'reports/some_report_template.html', context)

@login_required
def export_complaints_excel(request):
    complaints = Complaint.objects.all().order_by('complaint_no')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complaints"

    # Header row (without 'Actions')
    headers = [
        "No", "Date", "Center", "Area Manager", "Area Manager Email", "VIN", "Phone No",
        "Complaint No", "Immediate Action", "Correct Action", "Complaint Type",
        "Comment", "Status", "Reporter"
    ]
    ws.append(headers)

    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add complaint data
    for idx, complaint in enumerate(complaints, start=1):
        row = [
            idx,
            complaint.date.strftime('%Y-%m-%d'),
            complaint.center,
            complaint.area_manager,
            complaint.area_manager_email,
            complaint.vin,
            complaint.tele_no,
            complaint.complaint_no,
            complaint.immediate_action,
            complaint.correct_action,
            complaint.complaint_type,
            complaint.comment,
            complaint.status,
            complaint.reporter.username.replace(" ", "_") if complaint.reporter else '',
        ]
        ws.append(row)

    # Apply border and alignment to data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or '') for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5

    # Return Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formatted_complaints.xlsx"'
    wb.save(response)
    return response

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse

def export_classchange_excel(request):
    class_changes = ClassChange.objects.all()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Class Changes'

    # Define header row
    headers = [
        'No', 'Date', 'Time', 'Center', 'VIN', 'Previous Class', 'Change Class',
        'Reason', 'Approved By', 'Refund', 'Remark', 'Updated By'
    ]

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # nice blue color
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    even_row_fill = PatternFill("solid", fgColor="DCE6F1")  # light blue fill for even rows

    # Create header row with styles
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Fill data rows with border and alternating fills
    for idx, cc in enumerate(class_changes, start=1):
        row_num = idx + 1
        values = [
            idx,
            cc.date.strftime('%Y-%m-%d'),
            cc.time.strftime('%H:%M:%S'),
            cc.center,
            cc.vin,
            cc.previous_class,
            cc.change_class,
            cc.reason,
            cc.approved_by,
            'Yes' if cc.refund else 'No',
            cc.remark,
            cc.updated_by.username,
        ]

        for col_num, value in enumerate(values, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = even_row_fill

    # Adjust column widths automatically
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column].width = adjusted_width

    # Prepare response for download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=class_changes.xlsx'
    workbook.save(response)
    return response

